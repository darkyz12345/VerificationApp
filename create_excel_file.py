from pdf_extract_data import find_name_number_uncertainty
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.chart import LineChart, Reference


# PATH_TO_FILE_LAST_YEAR = 'Микрометр гладкий 19,01,23.pdf'
# PATH_TO_FILE_THIS_YEAR = 'Микрометр гладкий  13,04,24.pdf'
# PATH_TO_FILE_LAST_YEAR = 'WDW-100D  CY20230366.pdf'
# PATH_TO_FILE_THIS_YEAR = 'Цифровой_динамометрический_ключ_AWM_100.pdf'


def create_chart(row_last_year, unit, sheet):
    chart = LineChart()
    chart.title = f'Дрейф({row_last_year.values[0]} {unit})'
    # chart.style = 13
    data = Reference(sheet, min_col=4, min_row=7, max_col=7, max_row=9)
    categories = Reference(sheet, min_col=3, min_row=8, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    return chart

def create_excel(path_to_file_last_year, path_to_file_this_year, last_year, this_year, method_top, method_bottom, save_path):
    name_last_year, serie_number_last_year, uncertainty_last_year = find_name_number_uncertainty(path_to_file_last_year)
    # print(type(uncertainty_last_year))
    name_this_year, serie_number_this_year, uncertainty_this_year = find_name_number_uncertainty(path_to_file_this_year)
    # print(type(uncertainty_this_year))
    unit = uncertainty_this_year.columns[0].split(',')[-1].strip().split()[-1]
    for name_header in uncertainty_this_year.columns:
        if 'неопределенность' in str(name_header).lower() or 'u,' in str(name_header).lower():
            uncertainty_index = list(uncertainty_this_year).index(name_header)
            uncertainty_unit = name_header.split()[-1]
        if 'отклонение' in str(name_header).lower() or 'q,' in str(name_header).lower():
            deviation_index = list(uncertainty_this_year).index(name_header)
            deviation_unit = name_header.split()[-1]
    workbook = openpyxl.Workbook()
    for (index_last_year, row_last_year), (index_this_year, row_this_year) in zip(uncertainty_last_year.iterrows(),
                                                                                  uncertainty_this_year.iterrows()):
        tmp_func = lambda x: x.isdigit() or ',' == x or '.' == x
        if all(tmp_func(i) for i in str(row_last_year.values[0])) is False:
            continue
        sheet = workbook.create_sheet(title=f'point_{row_last_year.values[0]}')
        sheet['A1'] = 'Дата проведения верификации'
        sheet['A2'] = 'Марка, модель оборудования'
        sheet['A3'] = 'Заводской номер оборудования'
        sheet['A4'] = 'Результаты верификации'
        sheet['A5'] = 'Ответственный за верификацию оборудования'
        sheet['B1'] = datetime.now().strftime('%d.%m.%Y')
        sheet['B2'] = name_this_year
        sheet['B3'] = serie_number_this_year
        sheet['B5'] = 'Захарченко Е. Д.'
        sheet['C8'] = last_year
        sheet['C9'] = this_year
        sheet['D7'] = 'треб. метода'
        sheet['E7'] = 'треб. метода'
        sheet['F7'] = 'откл.'
        sheet['G7'] = 'неопр.'
        sheet['F6'] = unit
        sheet['G6'] = unit
        if '%' in uncertainty_unit or 'q' == uncertainty_unit:
            uncertainty_value_last_year = float(row_last_year.values[uncertainty_index].replace(',', '.')) * float(
                row_last_year.values[0].replace(',', '.')) / 100
            deviation_value_last_year = float(row_last_year.values[deviation_index].replace(',', '.')) * (
                    float(row_last_year.values[0].replace(',', '.')) / 100)
            uncertainty_value_this_year = float(row_this_year.values[uncertainty_index].replace(',', '.')) * float(
                row_this_year.values[0].replace(',', '.')) / 100
            deviation_value_this_year = (float(row_this_year.values[deviation_index].replace(',', '.')) *
                                         float(row_this_year.values[0].replace(',', '.')) / 100)
        else:
            uncertainty_value_last_year = row_last_year.values[uncertainty_index]
            deviation_value_last_year = row_last_year.values[deviation_index]
            uncertainty_value_this_year = row_this_year.values[uncertainty_index]
            deviation_value_this_year = row_this_year.values[deviation_index]
        sheet['F8'] = deviation_value_last_year
        sheet['F9'] = deviation_value_this_year
        sheet['G8'] = uncertainty_value_last_year
        sheet['G9'] = uncertainty_value_this_year
        sheet['D8'] = method_top
        sheet['D9'] = method_top
        sheet['E8'] = method_bottom
        sheet['E9'] = method_bottom
        dims = {}
        for row in sheet.rows:
            for cell in row:
                dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value + 5
        border_style = Border(left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              bottom=Side(border_style='thin', color='000000'))
        for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2):
            for cell in row:
                cell.border = border_style
        for row in sheet.iter_rows(min_row=7, max_row=9, min_col=3, max_col=7):
            for cell in row:
                cell.border = border_style
        sheet["F6"].border = border_style
        sheet["G6"].border = border_style
        sheet.add_chart(create_chart(row_last_year, unit, sheet))
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    workbook.save(f'{save_path}/{name_this_year}.xlsx')
    return f'{save_path}/{name_this_year}.xlsx'
